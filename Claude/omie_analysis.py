#!/usr/bin/env python3
"""
OMIE Energy Cost Comparison — Portugal 2025
============================================
Downloads OMIE day-ahead market prices for Portugal (Apr–Nov 2025),
merges them with 30-min household consumption data, and produces an
Excel report comparing what you paid on a standard (flat) tariff versus
what you would have paid on an OMIE-indexed tariff.

Usage:
    pip install requests pandas openpyxl
    python omie_analysis.py

Output:
    omie_analysis_output.xlsx  (saved in the same folder as this script)
"""

import requests
import pandas as pd
import time
import re
from datetime import date, timedelta
from pathlib import Path
from io import StringIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ─────────────────────────────────────────────
# CONFIGURATION — adjust as needed
# ─────────────────────────────────────────────

# Path to your Legrand consumption CSV
CSV_PATH = (
    Path.home()
    / "Library/Mobile Documents/com~apple~CloudDocs"
    / "2025 Legrand Apr - Nov Energy Exports"
    / "2025 Electricity Abril-Novembro.csv"
)

# Output file
OUTPUT_PATH = Path(__file__).parent / "omie_analysis_output.xlsx"

# Date range to analyse
START_DATE = date(2025, 4, 1)
END_DATE   = date(2025, 11, 30)

# OMIE-indexed tariff add-ons (edit to match your provider's offer)
OMIE_MARGIN_EUR_MWH  = 5.0    # Commercial margin charged on top of OMIE (€/MWh)
VAT_RATE             = 0.23   # IVA — 23% in Portugal
IEC_EUR_KWH          = 0.001  # Imposto Especial de Consumo (€/kWh, residential < 10 400 kWh/yr)
# Network / access tariff — regulated, identical on any tariff.
# Set to 0 if you only want to compare the energy+tax component.
# Or enter the per-kWh value from your OMIE provider's quote.
NETWORK_EUR_KWH      = 0.0    # e.g. 0.09 if you want a full bill estimate

# What you actually paid (standard tariff), from your electricity bills
ACTUAL_PAID = {
    "2025-04": {"kwh": 901,  "eur": 218.96},
    "2025-05": {"kwh": 730,  "eur": 177.50},
    "2025-06": {"kwh": 460,  "eur": 111.82},
    "2025-07": {"kwh": 816,  "eur": 198.36},
    "2025-08": {"kwh": 644,  "eur": 164.18},
    "2025-09": {"kwh": 489,  "eur": 124.74},
    "2025-10": {"kwh": 616,  "eur": 157.29},
    "2025-11": {"kwh": 1341, "eur": 342.07},
}

# ─────────────────────────────────────────────
# 1. DOWNLOAD OMIE PRICE FILES
# ─────────────────────────────────────────────

OMIE_URL  = "https://www.omie.es/pt/file-download?parents=marginalpdbcpt&filename=marginalpdbcpt_{date}.1"
CACHE_DIR = Path(__file__).parent / "_omie_cache"
CACHE_DIR.mkdir(exist_ok=True)

def download_omie_prices() -> pd.DataFrame:
    """Download (or read from cache) all daily OMIE price files and return a tidy DataFrame."""
    records = []
    current = START_DATE
    total = (END_DATE - START_DATE).days + 1
    print(f"Downloading {total} OMIE price files …")

    while current <= END_DATE:
        date_str = current.strftime("%Y%m%d")
        cache_file = CACHE_DIR / f"marginalpdbcpt_{date_str}.1"

        if cache_file.exists():
            raw = cache_file.read_text(encoding="latin-1")
        else:
            url = OMIE_URL.format(date=date_str)
            for attempt in range(3):
                try:
                    r = requests.get(url, timeout=30)
                    r.raise_for_status()
                    raw = r.text
                    cache_file.write_text(raw, encoding="latin-1")
                    break
                except Exception as e:
                    if attempt == 2:
                        print(f"  ✗ Failed {date_str}: {e}")
                        raw = None
                    else:
                        time.sleep(2)

        if raw:
            day_records = parse_omie_file(raw, current)
            records.extend(day_records)

        if len(records) % 240 == 0:  # progress every ~10 days
            days_done = (current - START_DATE).days + 1
            print(f"  {days_done}/{total} days processed …")

        current += timedelta(days=1)

    df = pd.DataFrame(records, columns=["datetime_hour", "omie_price_eur_mwh"])
    df["datetime_hour"] = pd.to_datetime(df["datetime_hour"])
    print(f"✓ OMIE prices loaded: {len(df)} hourly records\n")
    return df


def parse_omie_file(raw: str, day: date) -> list:
    """Parse one OMIE file and return list of (datetime, price_eur_mwh) tuples.

    Actual OMIE format (marginalpdbcpt):
      MARGINALPDBCPT;
      Year;Month;Day;Hour;PriceES;PricePT;
      e.g.  2025;04;01;1;90;90;

    Falls back to legacy 3-col format: DD/MM/YYYY;Hour;Price
    """
    records = []
    for line in raw.splitlines():
        line = line.strip().rstrip(";")   # remove trailing semicolon
        parts = [p.strip() for p in line.split(";")]

        try:
            # ── Primary format: Year;Month;Day;Hour;PriceES;PricePT (6 fields) ──
            if len(parts) == 6 and re.match(r"^\d{4}$", parts[0]):
                yr, mo, dy = int(parts[0]), int(parts[1]), int(parts[2])
                hour  = int(parts[3])
                if hour > 24:       # skip intraday slots (hours 25–96)
                    continue
                price = float(parts[5].replace(",", "."))  # col 5 = Portugal price
                dt = pd.Timestamp(year=yr, month=mo, day=dy, hour=0) + pd.Timedelta(hours=hour - 1)
                records.append((dt, price))
                continue

            # ── Fallback: Year;Month;Day;Hour;Price (5 fields, single price) ──
            if len(parts) == 5 and re.match(r"^\d{4}$", parts[0]):
                yr, mo, dy = int(parts[0]), int(parts[1]), int(parts[2])
                hour  = int(parts[3])
                if hour > 24:
                    continue
                price = float(parts[4].replace(",", "."))
                dt = pd.Timestamp(year=yr, month=mo, day=dy, hour=0) + pd.Timedelta(hours=hour - 1)
                records.append((dt, price))
                continue

            # ── Legacy: DD/MM/YYYY;Hour;Price (3 fields) ──
            if len(parts) >= 3 and re.match(r"\d{2}/\d{2}/\d{4}", parts[0]):
                hour  = int(parts[1])
                price = float(parts[2].replace(",", "."))
                dt = pd.Timestamp(day) + pd.Timedelta(hours=hour - 1)
                records.append((dt, price))

        except (ValueError, IndexError):
            continue

    return records


# ─────────────────────────────────────────────
# 2. LOAD CONSUMPTION CSV
# ─────────────────────────────────────────────

def load_consumption() -> pd.DataFrame:
    """Parse the Legrand 30-min consumption CSV and return a clean DataFrame."""
    # The CSV has a 6-line metadata header before the actual column headers
    raw = CSV_PATH.read_text(encoding="utf-8-sig")
    lines = raw.splitlines()

    # Find the data header row (contains 'Timestamp')
    header_row = None
    for i, line in enumerate(lines):
        if "Timestamp" in line:
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find 'Timestamp' header row in CSV")

    data_str = "\n".join(lines[header_row:])
    df = pd.read_csv(StringIO(data_str), header=0)

    # Rename for clarity — columns 0,1,2,3 are what we need
    cols = df.columns.tolist()
    df = df.rename(columns={
        cols[0]: "timestamp_unix",
        cols[1]: "datetime_str",
        cols[2]: "bill_eur",
        cols[3]: "consumption_wh",
    })

    df = df[["timestamp_unix", "datetime_str", "bill_eur", "consumption_wh"]].copy()
    df = df.dropna(subset=["timestamp_unix"])
    df["datetime_str"] = df["datetime_str"].astype(str).str.strip('"')

    # Parse the date string "2025/04/01 00:15:00" → datetime
    df["datetime"] = pd.to_datetime(df["datetime_str"], format="%Y/%m/%d %H:%M:%S", errors="coerce")
    df = df.dropna(subset=["datetime"])

    # Floor to hour so we can join with OMIE hourly prices
    df["datetime_hour"] = df["datetime"].dt.floor("h")

    df["bill_eur"]        = pd.to_numeric(df["bill_eur"],        errors="coerce").fillna(0)
    df["consumption_wh"]  = pd.to_numeric(df["consumption_wh"], errors="coerce").fillna(0)
    df["consumption_kwh"] = df["consumption_wh"] / 1000.0

    # Filter to our date range
    df = df[(df["datetime"] >= pd.Timestamp(START_DATE)) &
            (df["datetime"] <= pd.Timestamp(END_DATE) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1))]

    print(f"✓ Consumption data loaded: {len(df)} 30-min records\n")
    return df


# ─────────────────────────────────────────────
# 3. MERGE AND CALCULATE OMIE COST
# ─────────────────────────────────────────────

def build_analysis(consumption: pd.DataFrame, omie: pd.DataFrame) -> pd.DataFrame:
    """Merge consumption with OMIE prices and compute OMIE-indexed cost."""
    df = consumption.merge(omie, on="datetime_hour", how="left")

    missing = df["omie_price_eur_mwh"].isna().sum()
    if missing > 0:
        print(f"  ⚠ {missing} 30-min intervals have no OMIE price (filled with hourly average)")
        df["omie_price_eur_mwh"] = df.groupby(df["datetime"].dt.date)["omie_price_eur_mwh"].transform(
            lambda x: x.fillna(x.mean())
        )
        df["omie_price_eur_mwh"] = df["omie_price_eur_mwh"].fillna(df["omie_price_eur_mwh"].mean())

    # OMIE-indexed cost:
    #   energy component = (omie_price €/MWh + margin €/MWh) / 1000  →  €/kWh
    #   then add IEC + network (€/kWh), then multiply by (1 + VAT)
    df["omie_energy_eur_kwh"] = (df["omie_price_eur_mwh"] + OMIE_MARGIN_EUR_MWH) / 1000.0
    df["omie_cost_eur"] = (
        (df["omie_energy_eur_kwh"] + IEC_EUR_KWH + NETWORK_EUR_KWH) * (1 + VAT_RATE)
        * df["consumption_kwh"]
    )

    # Month label
    df["month"] = df["datetime"].dt.to_period("M").astype(str)

    return df


# ─────────────────────────────────────────────
# 4. MONTHLY SUMMARY
# ─────────────────────────────────────────────

def build_monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate to monthly level and join with actual paid data."""
    monthly = df.groupby("month").agg(
        consumption_kwh_csv=("consumption_kwh", "sum"),
        actual_cost_csv=("bill_eur", "sum"),
        omie_cost_eur=("omie_cost_eur", "sum"),
        omie_price_avg=("omie_price_eur_mwh", "mean"),
        omie_price_min=("omie_price_eur_mwh", "min"),
        omie_price_max=("omie_price_eur_mwh", "max"),
    ).reset_index()

    # Attach bill figures from the user-provided monthly totals
    bill_rows = []
    for month, info in ACTUAL_PAID.items():
        bill_rows.append({"month": month, "bill_kwh": info["kwh"], "bill_eur": info["eur"]})
    bills = pd.DataFrame(bill_rows)

    monthly = monthly.merge(bills, on="month", how="left")

    # Derived columns
    monthly["actual_rate_eur_kwh"]   = monthly["bill_eur"] / monthly["bill_kwh"]
    monthly["omie_rate_eur_kwh"]     = monthly["omie_cost_eur"] / monthly["consumption_kwh_csv"]
    monthly["saving_eur"]            = monthly["bill_eur"] - monthly["omie_cost_eur"]
    monthly["saving_pct"]            = monthly["saving_eur"] / monthly["bill_eur"] * 100

    # Month name for display
    month_names = {
        "2025-04": "Abril", "2025-05": "Maio", "2025-06": "Junho",
        "2025-07": "Julho", "2025-08": "Agosto", "2025-09": "Setembro",
        "2025-10": "Outubro", "2025-11": "Novembro",
    }
    monthly["month_name"] = monthly["month"].map(month_names).fillna(monthly["month"])

    return monthly


# ─────────────────────────────────────────────
# 5. BUILD EXCEL REPORT
# ─────────────────────────────────────────────

# Colour palette
COL_HEADER_BG = "1F4E79"   # dark blue
COL_HEADER_FG = "FFFFFF"
COL_ALT_ROW   = "D6E4F0"   # light blue
COL_ACTUAL    = "FF6B6B"   # coral
COL_OMIE      = "51CF66"   # green
COL_SAVING_POS = "51CF66"
COL_SAVING_NEG = "FF6B6B"

EURO_FMT = '#,##0.00 "€"'
KWH_FMT  = '#,##0.00 "kWh"'
PCT_FMT  = '0.0"%"'
RATE_FMT = '0.000 "€/kWh"'
MWH_FMT  = '#,##0.00 "€/MWh"'


def header_style(cell, text=None):
    if text is not None:
        cell.value = text
    cell.font = Font(name="Arial", bold=True, color=COL_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=COL_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def alt_row_fill(row_cells, row_idx):
    if row_idx % 2 == 0:
        fill = PatternFill("solid", fgColor=COL_ALT_ROW)
        for c in row_cells:
            c.fill = fill


def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_borders(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()


def write_summary_sheet(wb, monthly: pd.DataFrame):
    ws = wb.create_sheet("Resumo Mensal")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    for col in ["B","C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[col].width = 15

    # Title
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "Comparação Tarifário Fixo vs. OMIE Indexado — 2025 (Abril–Novembro)"
    title.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Config box
    ws.merge_cells("A2:K2")
    config_cell = ws["A2"]
    config_cell.value = (
        f"Parâmetros OMIE: Margem = {OMIE_MARGIN_EUR_MWH} €/MWh  |  "
        f"IVA = {VAT_RATE*100:.0f}%  |  IEC = {IEC_EUR_KWH*1000:.1f} €/MWh  |  "
        f"Rede/CIEG = {NETWORK_EUR_KWH*1000:.1f} €/MWh (se 0, só custo energia)"
    )
    config_cell.font = Font(name="Arial", italic=True, size=9, color="555555")
    config_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Column headers (row 4)
    headers = [
        "Mês",
        "Consumo\n(kWh)",
        "Pago\n(€)",
        "Taxa Fixa\n(€/kWh)",
        "Preço OMIE\nmédio (€/MWh)",
        "Preço OMIE\nmín (€/MWh)",
        "Preço OMIE\nmáx (€/MWh)",
        "Custo OMIE\nindexado (€)",
        "Taxa OMIE\n(€/kWh)",
        "Poupança\n(€)",
        "Poupança\n(%)",
    ]
    ws.row_dimensions[4].height = 40
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        header_style(cell, h)

    # Data rows (starting row 5)
    for row_idx, row in enumerate(monthly.itertuples(), start=1):
        excel_row = row_idx + 4
        ws.row_dimensions[excel_row].height = 18
        values = [
            row.month_name,
            row.bill_kwh,
            row.bill_eur,
            row.actual_rate_eur_kwh,
            row.omie_price_avg,
            row.omie_price_min,
            row.omie_price_max,
            row.omie_cost_eur,
            row.omie_rate_eur_kwh,
            row.saving_eur,
            row.saving_pct,
        ]
        fmts = [None, KWH_FMT, EURO_FMT, RATE_FMT, MWH_FMT, MWH_FMT, MWH_FMT, EURO_FMT, RATE_FMT, EURO_FMT, PCT_FMT]

        cells = []
        for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            if fmt:
                c.number_format = fmt
            # Colour savings column
            if col_idx == 10 and val is not None:  # saving_eur
                c.font = Font(name="Arial", size=10, bold=True,
                              color=COL_SAVING_POS if val >= 0 else COL_SAVING_NEG)
            if col_idx == 11 and val is not None:  # saving_pct
                c.font = Font(name="Arial", size=10, bold=True,
                              color=COL_SAVING_POS if val >= 0 else COL_SAVING_NEG)
            cells.append(c)

        alt_row_fill(cells, row_idx)

    # Totals row
    n_months = len(monthly)
    total_row = n_months + 5
    ws.row_dimensions[total_row].height = 20
    totals = [
        "TOTAL",
        monthly["bill_kwh"].sum(),
        monthly["bill_eur"].sum(),
        monthly["bill_eur"].sum() / monthly["bill_kwh"].sum(),
        monthly["omie_price_avg"].mean(),
        monthly["omie_price_min"].min(),
        monthly["omie_price_max"].max(),
        monthly["omie_cost_eur"].sum(),
        monthly["omie_cost_eur"].sum() / monthly["consumption_kwh_csv"].sum(),
        monthly["saving_eur"].sum(),
        monthly["saving_eur"].sum() / monthly["bill_eur"].sum() * 100,
    ]
    total_fmts = [None, KWH_FMT, EURO_FMT, RATE_FMT, MWH_FMT, MWH_FMT, MWH_FMT, EURO_FMT, RATE_FMT, EURO_FMT, PCT_FMT]
    for col_idx, (val, fmt) in enumerate(zip(totals, total_fmts), start=1):
        c = ws.cell(row=total_row, column=col_idx, value=val)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        if fmt:
            c.number_format = fmt

    apply_borders(ws, 4, total_row, 1, len(headers))

    # Note row
    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:K{note_row}")
    note = ws.cell(row=note_row, column=1)
    note.value = (
        "Nota: 'Custo OMIE indexado' = (Preço OMIE + Margem) × (1+IVA) × consumo + IEC + Rede. "
        "Para comparação completa da fatura, configure NETWORK_EUR_KWH no script com o valor do seu tarifário indexado."
    )
    note.font = Font(name="Arial", italic=True, size=8, color="777777")
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 30

    # ── Chart: actual vs OMIE monthly cost ──
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Custo Mensal: Tarifário Fixo vs. OMIE Indexado"
    chart.y_axis.title = "€"
    chart.x_axis.title = "Mês"
    chart.style = 10
    chart.width = 22
    chart.height = 12

    data_start_row = 5
    data_end_row   = 5 + n_months - 1

    # Actual paid (col 3)
    actual_ref = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)
    # OMIE cost (col 8)
    omie_ref   = Reference(ws, min_col=8, min_row=data_start_row, max_row=data_end_row)
    # Categories (month names, col 1)
    cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

    chart.add_data(actual_ref)
    chart.add_data(omie_ref)
    chart.set_categories(cats)
    chart.series[0].title.v = "Pago (€)"
    chart.series[1].title.v = "Custo OMIE indexado (€)"

    ws.add_chart(chart, f"A{note_row + 2}")
    return ws


def write_hourly_sheet(wb, df: pd.DataFrame):
    """Write the full hourly-level dataset to a separate sheet."""
    ws = wb.create_sheet("Dados Horários")
    ws.sheet_view.showGridLines = False

    hourly = df.groupby("datetime_hour").agg(
        consumption_kwh=("consumption_kwh", "sum"),
        actual_bill_eur=("bill_eur", "sum"),
        omie_cost_eur=("omie_cost_eur", "sum"),
        omie_price_eur_mwh=("omie_price_eur_mwh", "first"),
    ).reset_index()

    hourly["saving_eur"] = hourly["actual_bill_eur"] - hourly["omie_cost_eur"]
    hourly["month"] = hourly["datetime_hour"].dt.to_period("M").astype(str)

    headers = ["Data/Hora", "Consumo (kWh)", "Pago (€)", "Custo OMIE (€)", "Preço OMIE (€/MWh)", "Poupança (€)"]
    fmts    = [None, KWH_FMT, EURO_FMT, EURO_FMT, MWH_FMT, EURO_FMT]
    widths  = [20, 15, 13, 15, 18, 13]

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=1, column=i)
        header_style(cell, h)

    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(hourly.itertuples(), start=2):
        values = [row.datetime_hour, row.consumption_kwh, row.actual_bill_eur,
                  row.omie_cost_eur, row.omie_price_eur_mwh, row.saving_eur]
        cells = []
        for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
            if fmt:
                c.number_format = fmt
            if col_idx == 6 and val is not None:
                c.font = Font(name="Arial", size=9, color=COL_SAVING_POS if val >= 0 else COL_SAVING_NEG)
            cells.append(c)
        alt_row_fill(cells, row_idx)

    apply_borders(ws, 1, len(hourly) + 1, 1, len(headers))
    ws.freeze_panes = "A2"
    return ws


def write_heatmap_sheet(wb, df: pd.DataFrame):
    """Hourly average OMIE price heatmap by hour-of-day and month."""
    ws = wb.create_sheet("Heatmap Preços OMIE")
    ws.sheet_view.showGridLines = False

    df2 = df.copy()
    df2["hour"] = df2["datetime"].dt.hour
    df2["month"] = df2["datetime"].dt.to_period("M").astype(str)

    pivot = df2.groupby(["hour", "month"])["omie_price_eur_mwh"].mean().unstack("month")
    months_order = [f"2025-{m:02d}" for m in range(4, 12)]
    pivot = pivot.reindex(columns=[m for m in months_order if m in pivot.columns])

    month_names = {
        "2025-04": "Abr", "2025-05": "Mai", "2025-06": "Jun",
        "2025-07": "Jul", "2025-08": "Ago", "2025-09": "Set",
        "2025-10": "Out", "2025-11": "Nov",
    }

    title = ws.cell(row=1, column=1, value="Preço Médio OMIE por Hora e Mês (€/MWh)")
    title.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws.merge_cells(f"A1:{get_column_letter(len(pivot.columns)+1)}1")
    ws.row_dimensions[1].height = 24

    # Column headers
    ws.cell(row=2, column=1, value="Hora").font = Font(name="Arial", bold=True)
    for col_idx, month in enumerate(pivot.columns, start=2):
        c = ws.cell(row=2, column=col_idx, value=month_names.get(month, month))
        header_style(c)
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    ws.column_dimensions["A"].width = 8

    # Find global min/max for colour scale
    vmin = pivot.min().min()
    vmax = pivot.max().max()

    def price_to_color(val):
        if pd.isna(val):
            return "FFFFFF"
        ratio = (val - vmin) / max(vmax - vmin, 0.01)
        # Green (low price) → Yellow → Red (high price)
        if ratio < 0.5:
            r = int(255 * ratio * 2)
            g = 200
        else:
            r = 220
            g = int(200 * (1 - (ratio - 0.5) * 2))
        return f"{r:02X}{g:02X}50"

    for hour in range(24):
        ws.cell(row=hour + 3, column=1, value=f"{hour:02d}:00").font = Font(name="Arial", size=9)
        ws.row_dimensions[hour + 3].height = 16
        for col_idx, month in enumerate(pivot.columns, start=2):
            val = pivot.loc[hour, month] if hour in pivot.index else None
            c = ws.cell(row=hour + 3, column=col_idx)
            if val is not None and not pd.isna(val):
                c.value = round(val, 1)
                c.fill = PatternFill("solid", fgColor=price_to_color(val))
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center")
            c.number_format = "0.0"

    return ws


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print("=" * 60)
    print(" OMIE Energy Cost Comparison — Portugal 2025")
    print("=" * 60)

    # Step 1: download OMIE prices
    omie = download_omie_prices()

    # Step 2: load consumption data
    consumption = load_consumption()

    # Step 3: merge and compute
    df = build_analysis(consumption, omie)

    # Step 4: monthly summary
    monthly = build_monthly_summary(df)

    print("\nMonthly summary:")
    print(monthly[["month_name", "bill_eur", "omie_cost_eur", "saving_eur", "saving_pct"]].to_string(index=False))
    total_saving = monthly["saving_eur"].sum()
    print(f"\nTotal saving Apr–Nov 2025: {total_saving:+.2f} €  "
          f"({'cheaper on OMIE' if total_saving > 0 else 'more expensive on OMIE'})")

    # Step 5: write Excel
    print("\nBuilding Excel report …")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_summary_sheet(wb, monthly)
    write_hourly_sheet(wb, df)
    write_heatmap_sheet(wb, df)

    wb.save(OUTPUT_PATH)
    print(f"\n✓ Report saved to: {OUTPUT_PATH}\n")


if __name__ == "__main__":
    main()
